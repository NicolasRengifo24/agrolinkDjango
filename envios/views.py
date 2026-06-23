import csv
import io

from django.shortcuts import render , redirect, get_object_or_404
from django.http import JsonResponse
from django.core.paginator import Paginator
from django.db.models import Q, Sum, Count
from django.db.models.functions import TruncMonth
from .models import Envio , Vehiculo
from productos.models import Producto, Finca, CategoriaProducto
from usuarios.models import Usuario, Transportista, Administrador, Notificacion
from django.contrib.auth.decorators import login_required
from django.contrib import messages 

import json
from datetime import datetime
import random
import string

# Create your views here.
@login_required
def inicio_transportista(request):
    
    envios = Envio.objects.select_related(
        "id_compra__id_cliente",
        "id_vehiculo",
        "id_transportista"
    ).filter(id_transportista__isnull=True,estado_envio="pendiente").order_by('-id_envio')
    
    # Obtener vehículos activos del transportista actual
    vehiculos_activos = []
    if request.user.is_authenticated:
        try:
            usuario_obj = Usuario.objects.get(user=request.user)
            transportista_obj = Transportista.objects.get(id_usuario=usuario_obj)
            vehiculos_activos = Vehiculo.objects.filter(
                id_transportista=transportista_obj,
                estado='ACTIVO'
            )
        except (Usuario.DoesNotExist, Transportista.DoesNotExist):
            pass
    
    # ── FILTROS ──
    ciudad_origen = request.GET.get('ciudad_origen', '').strip()
    ciudad_destino = request.GET.get('ciudad_destino', '').strip()
    categoria = request.GET.get('categoria', '').strip()
    peso_max = request.GET.get('peso_max', '').strip()

    if ciudad_origen:
        envios = envios.filter(
            id_compra__detallescompra__id_producto__fincas__id_finca__ciudad__icontains=ciudad_origen
        )
    if ciudad_destino:
        envios = envios.filter(
            id_compra__id_cliente__id_usuario__ciudad__icontains=ciudad_destino
        )
    if categoria:
        envios = envios.filter(
            id_compra__detallescompra__id_producto__id_categoria_id=categoria
        )
    if peso_max:
        envios = envios.filter(peso_total_kg__lte=float(peso_max))

    envios = envios.distinct()

    # ── PAGINACIÓN ──
    pagina = request.GET.get('page', 1)
    paginator = Paginator(envios, 5)
    page_obj = paginator.get_page(pagina)

    # ── DATOS JSON PARA MAPAS (solo página actual) ──
    data_envios = []
    for envio in page_obj:
        lat_origen = None
        lng_origen = None
        direccion_origen = ""
        nombre_finca = ""
        lat_destino = None
        lng_destino = None
        direccion_destino = envio.id_compra.direccion_entrega or ""

        if envio.id_compra.latitud_destino and envio.id_compra.longitud_destino:
            lat_destino = float(envio.id_compra.latitud_destino)
            lng_destino = float(envio.id_compra.longitud_destino)

        detalles = envio.id_compra.detallescompra_set.all()
        if detalles.exists():
            primer_detalle = detalles.first()
            producto = primer_detalle.id_producto
            producto_finca_rel = producto.fincas.first()
            if producto_finca_rel:
                finca = producto_finca_rel.id_finca
                if finca:
                    nombre_finca = finca.nombre_finca or ""
                    direccion_origen = finca.direccion_finca or nombre_finca
                    if finca.latitud and finca.longitud:
                        lat_origen = float(finca.latitud)
                        lng_origen = float(finca.longitud)

        if lat_origen and lng_origen and lat_destino and lng_destino:
            data_envios.append({
                "id": envio.id_envio,
                "origen": [lat_origen, lng_origen],
                "destino": [lat_destino, lng_destino],
                "numero": envio.numero_seguimiento or f"ENV-{envio.id_envio}",
                "direccion_origen": direccion_origen,
                "direccion_destino": direccion_destino,
                "nombre_finca": nombre_finca,
                "peso": float(envio.peso_total_kg or 0),
                "distancia": float(envio.distancia_km or 0)
            })

    envios_json = json.dumps(data_envios)

    # ── DATOS PARA DROPDOWNS ──
    ciudades_origen = (
        Finca.objects.exclude(ciudad__isnull=True).exclude(ciudad='')
        .values_list('ciudad', flat=True).distinct().order_by('ciudad')
    )
    ciudades_destino = (
        Usuario.objects.filter(rol='Cliente')
        .exclude(ciudad__isnull=True).exclude(ciudad='')
        .values_list('ciudad', flat=True).distinct().order_by('ciudad')
    )
    categorias = CategoriaProducto.objects.all().order_by('nombre_categoria')

    return render(request, 'envios/envios_dashboard.html', {
        'page_obj': page_obj,
        'envios_json': envios_json,
        'vehiculos_activos': vehiculos_activos,
        'filtro_ciudad_origen': ciudad_origen,
        'filtro_ciudad_destino': ciudad_destino,
        'filtro_categoria': categoria,
        'filtro_peso_max': peso_max,
        'ciudades_origen': ciudades_origen,
        'ciudades_destino': ciudades_destino,
        'categorias': categorias,
        'kpi_disponibles': page_obj.paginator.count,
    })
   

@login_required
def mostrar_vehiculos(request):
    try:
        usuario_obj = Usuario.objects.get(user=request.user)
        transportista_obj = Transportista.objects.get(id_usuario=usuario_obj)
        vehiculos = Vehiculo.objects.filter(id_transportista=transportista_obj)

        notificaciones = Notificacion.objects.filter(
            destino=usuario_obj,
            tipo__in=['APROBACION_VEHICULO', 'RECHAZO_VEHICULO'],
            leido=False
        )
        for n in notificaciones:
            if n.tipo == 'APROBACION_VEHICULO':
                messages.success(request, n.mensaje)
            elif n.tipo == 'RECHAZO_VEHICULO':
                messages.error(request, n.mensaje)

        notificaciones_no_leidas = notificaciones.count()

    except (Usuario.DoesNotExist, Transportista.DoesNotExist):
        vehiculos = Vehiculo.objects.none()
        notificaciones_no_leidas = 0
        messages.warning(request, "No tienes perfil de transportista")
    
    return render(request, 'vehiculos/vehiculos_dashboard.html', {
        'vehiculos': vehiculos,
        'notificaciones_no_leidas': notificaciones_no_leidas,
    })
    
    
@login_required
def agregar_vehiculo(request):
    if request.method == 'POST':
        try:
            usuario_obj = Usuario.objects.get(user=request.user)
            transportista_obj = Transportista.objects.get(id_usuario=usuario_obj)
        except (Usuario.DoesNotExist, Transportista.DoesNotExist):
            messages.error(request, "No tienes permisos de transportista")
            return redirect('mostrar_vehiculos')

        tipo_vehiculo = request.POST.get('tipo_vehiculo')
        placa_vehiculo = request.POST.get('placa_vehiculo')
        capacidad_carga = request.POST.get('capacidad_carga')
        archivo = request.FILES.get('documento_propiedad')  # 👈 aquí

        if not tipo_vehiculo or not placa_vehiculo:
            messages.error(request, "Por favor completa los campos obligatorios")
            return redirect('mostrar_vehiculos')

        try:
            vehiculo = Vehiculo.objects.create(
                id_transportista=transportista_obj,
                tipo_vehiculo=tipo_vehiculo,
                placa_vehiculo=placa_vehiculo.upper(),
                capacidad_carga=capacidad_carga if capacidad_carga else 0,
                documento_propiedad=archivo,
                estado='PENDIENTE'
            )

            admins = Administrador.objects.select_related('id_usuario').all()
            for admin in admins:
                Notificacion.objects.create(
                    tipo='SOLICITUD_VEHICULO',
                    mensaje=f"El transportista {usuario_obj.nombre} {usuario_obj.apellido} ha registrado el vehículo {placa_vehiculo.upper()} y está pendiente de aprobación.",
                    destino=admin.id_usuario,
                    id_vehiculo=vehiculo
                )

            messages.success(request, f"Vehículo {placa_vehiculo} registrado. Pendiente de aprobación por un administrador.")
        except Exception as e:
            messages.error(request, f"Error al registrar vehículo: {str(e)}")

        return redirect('mostrar_vehiculos')

    return redirect('mostrar_vehiculos')


@login_required
def editar_vehiculo(request, vehiculo_id):
    try:
        usuario_obj = Usuario.objects.get(user=request.user)
        transportista_obj = Transportista.objects.get(id_usuario=usuario_obj)
    except (Usuario.DoesNotExist, Transportista.DoesNotExist):
        messages.error(request, "No tienes permisos de transportista")
        return redirect('mostrar_vehiculos')

    vehiculo = get_object_or_404(
        Vehiculo,
        id_vehiculo=vehiculo_id,
        id_transportista=transportista_obj
    )

    if vehiculo.estado not in ('PENDIENTE', 'RECHAZADO'):
        messages.error(request, "No puedes editar un vehículo que ya fue aprobado.")
        return redirect('mostrar_vehiculos')

    if request.method == 'POST':
        tipo_vehiculo = request.POST.get('tipo_vehiculo')
        placa_vehiculo = request.POST.get('placa_vehiculo')
        capacidad_carga = request.POST.get('capacidad_carga')
        archivo = request.FILES.get('documento_propiedad')

        if not tipo_vehiculo or not placa_vehiculo:
            messages.error(request, "Por favor completa los campos obligatorios")
            return render(request, 'vehiculos/editar_vehiculo.html', {'vehiculo': vehiculo})

        try:
            vehiculo.tipo_vehiculo = tipo_vehiculo
            vehiculo.placa_vehiculo = placa_vehiculo.upper()
            vehiculo.capacidad_carga = capacidad_carga if capacidad_carga else 0
            if archivo:
                vehiculo.documento_propiedad = archivo
            if vehiculo.estado == 'RECHAZADO':
                vehiculo.estado = 'PENDIENTE'

                admins = Administrador.objects.select_related('id_usuario').all()
                for admin in admins:
                    Notificacion.objects.create(
                        tipo='SOLICITUD_VEHICULO',
                        mensaje=f"El transportista {usuario_obj.nombre} {usuario_obj.apellido} ha editado el vehículo {placa_vehiculo.upper()} y está pendiente de re-aprobación.",
                        destino=admin.id_usuario,
                        id_vehiculo=vehiculo
                    )

            vehiculo.save()

            messages.success(request, f"Vehículo {placa_vehiculo} actualizado correctamente.")
        except Exception as e:
            messages.error(request, f"Error al actualizar vehículo: {str(e)}")

        return redirect('mostrar_vehiculos')

    return render(request, 'vehiculos/editar_vehiculo.html', {'vehiculo': vehiculo})


@login_required
def cambiar_estado_vehiculo(request, vehiculo_id):
    """Cambia el estado del vehículo (AJAX)"""
    
    if request.method == 'POST':
        try:
            usuario_obj = Usuario.objects.get(user=request.user)
            transportista_obj = Transportista.objects.get(id_usuario=usuario_obj)
            
            vehiculo = get_object_or_404(
                Vehiculo, 
                id_vehiculo=vehiculo_id, 
                id_transportista=transportista_obj
            )

            if vehiculo.estado in ('PENDIENTE', 'RECHAZADO'):
                messages.error(request, "No puedes cambiar el estado de un vehículo que no ha sido aprobado por el administrador.")
                return redirect('mostrar_vehiculos')
            
            # Cambiar estado
            if vehiculo.estado == 'ACTIVO':
                vehiculo.estado = 'SUSPENDIDO'
                mensaje = f"Vehículo {vehiculo.placa_vehiculo} suspendido"
            else:
                vehiculo.estado = 'ACTIVO'
                mensaje = f"Vehículo {vehiculo.placa_vehiculo} activado"
            
            vehiculo.save()
            
            return redirect('mostrar_vehiculos')
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'message': 'Método no permitido'})


@login_required
def eliminar_vehiculo(request, vehiculo_id):
    """Elimina un vehículo"""
    
    if request.method == 'POST':
        try:
            # Verificar el usuario
            usuario_obj = Usuario.objects.get(user=request.user)
            transportista_obj = Transportista.objects.get(id_usuario=usuario_obj)
            
            # Buscar el vehículo y verificar que pertenece al transportista
            vehiculo = get_object_or_404(
                Vehiculo, 
                id_vehiculo=vehiculo_id, 
                id_transportista=transportista_obj
            )
            
            placa = vehiculo.placa_vehiculo
            vehiculo.delete()
            
            messages.success(request, f"Vehículo {placa} eliminado correctamente")
            
        except Usuario.DoesNotExist:
            messages.error(request, "Perfil de usuario no encontrado")
        except Transportista.DoesNotExist:
            messages.error(request, "No tienes permisos de transportista")
        except Exception as e:
            messages.error(request, f"Error al eliminar: {str(e)}")
    
    return redirect('mostrar_vehiculos')


@login_required
def notificaciones_transportista(request):
    try:
        usuario_obj = Usuario.objects.get(user=request.user)
        notificaciones = Notificacion.objects.filter(
            destino=usuario_obj,
            tipo__in=['APROBACION_VEHICULO', 'RECHAZO_VEHICULO']
        ).order_by('-fecha_creacion')

        no_leidas = notificaciones.filter(leido=False).count()
    except Usuario.DoesNotExist:
        notificaciones = Notificacion.objects.none()
        no_leidas = 0

    return render(request, 'vehiculos/notificaciones.html', {
        'notificaciones': notificaciones,
        'no_leidas': no_leidas,
    })


@login_required
def marcar_notif_transportista(request, notif_id):
    notificacion = get_object_or_404(
        Notificacion, id_notificacion=notif_id, destino__user=request.user
    )
    notificacion.leido = True
    notificacion.save()
    return redirect('notificaciones_transportista')


@login_required
def mis_envios(request):

    try:
        usuario_obj = Usuario.objects.get(user=request.user)
        transportista_obj = Transportista.objects.get(id_usuario=usuario_obj)
    except Usuario.DoesNotExist:
        messages.error(request, "Usuario no encontrado")
        return redirect('inicio')
    except Transportista.DoesNotExist:
        messages.error(request, "No tienes perfil de transportista")
        return redirect('inicio')

    envios = Envio.objects.select_related(
        'id_compra__id_cliente__id_usuario',
        'id_vehiculo'
    ).filter(
        id_transportista=transportista_obj
    ).order_by('-id_envio')

    # ── Datos para modales de detalle ──
    envios_detalle = []
    for envio in envios:
        detalle = envio.id_compra.detallescompra_set.first() if envio.id_compra else None
        producto = detalle.id_producto if detalle else None
        productor = producto.id_usuario if producto else None
        finca_rel = producto.fincas.first() if producto else None
        finca = finca_rel.id_finca if finca_rel else None

        envios_detalle.append({
            'id_envio': envio.id_envio,
            'finca_nombre': finca.nombre_finca if finca else 'No especificada',
            'finca_direccion': finca.direccion_finca if finca else '',
            'finca_ciudad': finca.ciudad if finca else '',
            'productor_nombre': f"{productor.id_usuario.nombre} {productor.id_usuario.apellido}" if productor else 'No especificado',
            'productor_telefono': productor.id_usuario.telefono if productor else '',
            'cliente_nombre': f"{envio.id_compra.id_cliente.id_usuario.nombre} {envio.id_compra.id_cliente.id_usuario.apellido}" if envio.id_compra else 'N/A',
            'cliente_telefono': envio.id_compra.id_cliente.id_usuario.telefono if envio.id_compra else '',
            'cliente_direccion': envio.direccion_destino or '',
            'producto_nombre': producto.nombre_producto if producto else '',
            'foto_carga_url': envio.foto_carga.url if envio.foto_carga else '',
            'foto_descarga_url': envio.foto_descarga.url if envio.foto_descarga else '',
        })

    envios_detalle_json = json.dumps(envios_detalle)

    # ── Conteo de fotos pendientes ──
    fotos_pendientes = envios.filter(
        Q(estado_envio='En_Transito', foto_carga__isnull=True) |
        Q(estado_envio='Entregado', foto_carga__isnull=True) |
        Q(estado_envio='Entregado', foto_descarga__isnull=True)
    ).count()

    return render(request, 'envios/mis_envios_dashboard.html', {
        'envios': envios,
        'envios_detalle_json': envios_detalle_json,
        'fotos_pendientes': fotos_pendientes,
    })
    
    
def _fotos_pendientes_data(transportista_obj):
    """Devuelve lista de envíos del transportista con fotos faltantes."""
    qs = Envio.objects.filter(id_transportista=transportista_obj).exclude(
        estado_envio__in=['pendiente', 'Cancelado', 'cancelado']
    ).filter(
        Q(foto_carga__isnull=True) | Q(foto_descarga__isnull=True)
    )
    data = []
    for e in qs:
        faltantes = []
        if not e.foto_carga:
            faltantes.append('Carga')
        if not e.foto_descarga:
            faltantes.append('Descarga')
        detalle = e.id_compra.detallescompra_set.first() if e.id_compra else None
        producto = detalle.id_producto.nombre_producto if detalle else '—'
        data.append({
            'id': e.id_envio,
            'numero_seguimiento': e.numero_seguimiento or f'#{e.id_envio}',
            'producto': producto,
            'faltantes': faltantes,
        })
    return data


@login_required
def verificar_fotos_pendientes(request):
    """GET — Verifica si el transportista tiene envíos con fotos pendientes."""
    try:
        usuario_obj = Usuario.objects.get(user=request.user)
        transportista_obj = Transportista.objects.get(id_usuario=usuario_obj)
        data = _fotos_pendientes_data(transportista_obj)
        return JsonResponse({'pendientes': len(data) > 0, 'envios': data})
    except Exception:
        return JsonResponse({'pendientes': False, 'envios': []})


@login_required
def aceptar_viaje(request, envio_id):
    """Acepta un viaje y asigna vehículo, fechas y número de seguimiento"""
    
    if request.method == 'POST':
        try:
            # Obtener el transportista actual
            usuario_obj = Usuario.objects.get(user=request.user)
            transportista_obj = Transportista.objects.get(id_usuario=usuario_obj)

            # ── Validar fotos pendientes ──
            fotos_pend = _fotos_pendientes_data(transportista_obj)
            if fotos_pend:
                return JsonResponse({
                    'success': False,
                    'fotos_pendientes': True,
                    'envios': fotos_pend,
                    'message': 'Tienes fotos pendientes por subir'
                }, status=400)
            
            # Obtener el envío
            envio = get_object_or_404(Envio, id_envio=envio_id)
            
            # Verificar que el envío esté disponible (no asignado)
            if envio.id_transportista:
                return JsonResponse({
                    'success': False,
                    'message': 'Este envío ya fue asignado a otro transportista'
                }, status=400)
            
            # Obtener datos del formulario
            vehiculo_id = request.POST.get('vehiculo_id')
            fecha_recoleccion = request.POST.get('fecha_recoleccion')
            fecha_entrega_estimada = request.POST.get('fecha_entrega_estimada')
            
            # Validar que se seleccionó un vehículo
            if not vehiculo_id:
                return JsonResponse({
                    'success': False,
                    'message': 'Debes seleccionar un vehículo'
                }, status=400)
            
            # Obtener el vehículo y verificar que pertenece al transportista
            vehiculo = get_object_or_404(
                Vehiculo, 
                id_vehiculo=vehiculo_id, 
                id_transportista=transportista_obj
            )
            
            # Validar que el vehículo no esté suspendido
            if vehiculo.estado != 'ACTIVO':
                return JsonResponse({
                    'success': False,
                    'message': f'El vehículo {vehiculo.placa_vehiculo} está suspendido. No puede ser asignado.'
                }, status=400)
            
            # Validar fechas
            if not fecha_recoleccion or not fecha_entrega_estimada:
                return JsonResponse({
                    'success': False,
                    'message': 'Debes seleccionar ambas fechas'
                }, status=400)
            
            # Generar número de seguimiento único
            numero_seguimiento = generar_numero_seguimiento()
            
            # Calcular costo total
            distancia_km = envio.distancia_km or 0
            peso_total_kg = envio.peso_total_kg or 0
            
            # Tarifas
            tarifa_por_km = 3000  # $3,000 por km
            tarifa_por_kg = 200   # $200 por kg
            
            costo_distancia = distancia_km * tarifa_por_km
            costo_peso = peso_total_kg * tarifa_por_kg
            costo_total = costo_distancia + costo_peso
            
            # Actualizar el envío
            envio.id_transportista = transportista_obj
            envio.id_vehiculo = vehiculo
            envio.estado_envio = 'Asignado'
            envio.fecha_salida = fecha_recoleccion
            envio.fecha_entrega = fecha_entrega_estimada
            envio.numero_seguimiento = numero_seguimiento
            envio.tarifa_por_km = tarifa_por_km
            envio.tarifa_por_kg = tarifa_por_kg
            envio.costo_total = costo_total
            envio.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Viaje aceptado exitosamente. Número de seguimiento: {numero_seguimiento}',
                'numero_seguimiento': numero_seguimiento,
                'costo_total': f'{costo_total:,.0f}'
            })
            
        except Vehiculo.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Vehículo no encontrado o no te pertenece'
            }, status=400)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al aceptar el viaje: {str(e)}'
            }, status=400)
    
    return JsonResponse({'success': False, 'message': 'Método no permitido'}, status=405)

def generar_numero_seguimiento():
    """Genera un número de seguimiento único"""
    while True:
        # Formato: AGRO-YYYYMMDD-XXXXX
        fecha_actual = datetime.now().strftime('%Y%m%d')
        random_chars = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
        numero = f"AGRO-{fecha_actual}-{random_chars}"
        
        # Verificar que no exista
        if not Envio.objects.filter(numero_seguimiento=numero).exists():
            return numero
        
        
@login_required
def cambiar_estado_envio(request, envio_id, nuevo_estado):

    try:
        usuario_obj = Usuario.objects.get(user=request.user)
        transportista_obj = Transportista.objects.get(id_usuario=usuario_obj)

        envio = get_object_or_404(Envio, id_envio=envio_id)

        if envio.id_transportista != transportista_obj:
            messages.error(request, "No tienes permiso para este envío")
            return redirect('mis_envios')

        if envio.estado_envio == "Asignado" and nuevo_estado == "En_Transito":
            envio.estado_envio = "En_Transito"
            foto = request.FILES.get('foto_carga')
            if foto:
                envio.foto_carga = foto

        elif envio.estado_envio == "En_Transito" and nuevo_estado == "Entregado":
            envio.estado_envio = "Entregado"
            foto = request.FILES.get('foto_descarga')
            if foto:
                envio.foto_descarga = foto

        else:
            messages.warning(request, "Cambio de estado no permitido")
            return redirect('mis_envios')

        envio.save()
        messages.success(request, f"Estado actualizado a {envio.estado_envio}")

    except Exception as e:
        messages.error(request, f"Error: {str(e)}")

    return redirect('mis_envios')


@login_required
def subir_foto_envio(request, envio_id):
    """Endpoint para que el transportista suba fotos pendientes después."""

    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Método no permitido'}, status=405)

    try:
        usuario_obj = Usuario.objects.get(user=request.user)
        transportista_obj = Transportista.objects.get(id_usuario=usuario_obj)
        envio = get_object_or_404(Envio, id_envio=envio_id)

        if envio.id_transportista != transportista_obj:
            return JsonResponse({'success': False, 'message': 'No tienes permiso'}, status=403)

        tipo = request.POST.get('tipo')
        foto = request.FILES.get('foto')

        if not tipo or not foto:
            return JsonResponse({'success': False, 'message': 'Faltan tipo o archivo'}, status=400)

        if tipo == 'foto_carga':
            envio.foto_carga = foto
        elif tipo == 'foto_descarga':
            envio.foto_descarga = foto
        else:
            return JsonResponse({'success': False, 'message': 'Tipo inválido'}, status=400)

        envio.save()
        return JsonResponse({'success': True, 'message': 'Foto subida correctamente'})

    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


######### Esto es para el Api #######
from rest_framework.decorators import api_view
from rest_framework.response import Response
from .services import UbicacionService

@api_view(['GET'])
def ciudades_cundinamarca_api(request):
    ciudades = UbicacionService.obtener_ciudades_cundinamarca()
    return Response({
        "success": True,
        "data": ciudades
    })
    
    
## carga masiva 
@login_required
def cargar_vehiculos_csv(request):
    """Carga masiva de vehículos desde archivo CSV"""
    
    if request.method == 'POST':
        archivo = request.FILES.get('archivo')
        
        if not archivo:
            messages.error(request, 'Debes seleccionar un archivo CSV')
            return redirect('mostrar_vehiculos')
        
        if not archivo.name.endswith('.csv'):
            messages.error(request, 'El archivo debe ser de tipo CSV')
            return redirect('mostrar_vehiculos')
        
        # Obtener el transportista
        try:
            from usuarios.models import Usuario, Transportista
            usuario = Usuario.objects.get(user=request.user)
            transportista = Transportista.objects.get(id_usuario=usuario)
            print(f"Transportista ID: {transportista.id_usuario.id_usuario}")
        except Usuario.DoesNotExist:
            messages.error(request, 'No tienes un perfil de usuario asociado')
            return redirect('mostrar_vehiculos')
        except Transportista.DoesNotExist:
            messages.error(request, 'No tienes un perfil de transportista asociado')
            return redirect('mostrar_vehiculos')
        
        # Leer el archivo
        contenido = None
        for codificacion in ['utf-8', 'latin-1', 'iso-8859-1', 'windows-1252']:
            try:
                archivo.seek(0)
                contenido = archivo.read().decode(codificacion)
                break
            except UnicodeDecodeError:
                continue
        
        if not contenido:
            messages.error(request, 'No se pudo leer el archivo')
            return redirect('mostrar_vehiculos')
        
        # Detectar delimitador
        primera_linea = contenido.splitlines()[0]
        if '\t' in primera_linea:
            delimiter = '\t'
        elif ';' in primera_linea:
            delimiter = ';'
        else:
            delimiter = ','
        
        print(f"Delimitador detectado: '{delimiter}'")
        
        try:
            import csv
            import io
            
            io_string = io.StringIO(contenido)
            reader = csv.DictReader(io_string, delimiter=delimiter)
            
            vehiculos_creados = 0
            errores = []
            fila_num = 0
            
            for row in reader:
                fila_num += 1
                
                if not any(row.values()):
                    print(f"Fila {fila_num}: VACÍA")
                    continue
                
                print(f"\n--- Procesando fila {fila_num} ---")
                print(f"Datos: {dict(row)}")
                
                try:
                    tipo_vehiculo = row.get('tipo_vehiculo', '').strip()
                    capacidad_carga = row.get('capacidad_carga', 0)
                    placa_vehiculo = row.get('placa_vehiculo', '').strip().upper()
                    estado = row.get('estado', 'ACTIVO').strip().upper()
                    documento_propiedad = row.get('documento_propiedad', '').strip()
                    
                    print(f"Tipo: '{tipo_vehiculo}'")
                    print(f"Capacidad: '{capacidad_carga}'")
                    print(f"Placa: '{placa_vehiculo}'")
                    print(f"Estado: '{estado}'")
                    
                    # Validación 1: Tipo de vehículo
                    if not tipo_vehiculo:
                        error_msg = f"Fila {fila_num}: Tipo de vehículo requerido"
                        errores.append(error_msg)
                        print(f"❌ {error_msg}")
                        continue
                    
                    # Validación 2: Placa
                    if not placa_vehiculo:
                        error_msg = f"Fila {fila_num}: Placa requerida"
                        errores.append(error_msg)
                        print(f"❌ {error_msg}")
                        continue
                    
                    # Validación 3: Capacidad
                    try:
                        capacidad_float = float(capacidad_carga)
                        if capacidad_float <= 0:
                            error_msg = f"Fila {fila_num}: Capacidad de carga debe ser mayor a 0"
                            errores.append(error_msg)
                            print(f"❌ {error_msg}")
                            continue
                    except ValueError:
                        error_msg = f"Fila {fila_num}: Capacidad de carga debe ser un número"
                        errores.append(error_msg)
                        print(f"❌ {error_msg}")
                        continue
                    
                    # Validación 4: Placa duplicada
                    if Vehiculo.objects.filter(placa_vehiculo=placa_vehiculo).exists():
                        error_msg = f"Fila {fila_num}: La placa {placa_vehiculo} YA EXISTE en la base de datos"
                        errores.append(error_msg)
                        print(f"❌ {error_msg}")
                        # Mostrar el vehículo existente
                        vehiculo_existente = Vehiculo.objects.get(placa_vehiculo=placa_vehiculo)
                        print(f"   Vehículo existente ID: {vehiculo_existente.id_vehiculo}, Transportista: {vehiculo_existente.id_transportista.id_usuario.nombre_usuario}")
                        continue
                    
                    # Normalizar tipo
                    tipo_original = tipo_vehiculo
                    if tipo_vehiculo.lower() == 'camion':
                        tipo_vehiculo = 'Camión'
                    
                    tipos_validos = ['Camión', 'Camioneta', 'Van', 'Furgón', 'Volqueta']
                    if tipo_vehiculo not in tipos_validos:
                        error_msg = f"Fila {fila_num}: Tipo '{tipo_vehiculo}' no válido. Válidos: {', '.join(tipos_validos)}"
                        errores.append(error_msg)
                        print(f"❌ {error_msg}")
                        continue
                    
                    # Validar estado
                    if estado not in ['ACTIVO', 'SUSPENDIDO', 'MANTENIMIENTO']:
                        estado = 'ACTIVO'
                    
                    # Crear vehículo
                    print(f"Creando vehículo con:")
                    print(f"  - transportista: {transportista}")
                    print(f"  - tipo: {tipo_vehiculo}")
                    print(f"  - capacidad: {capacidad_float}")
                    print(f"  - placa: {placa_vehiculo}")
                    print(f"  - estado: {estado}")
                    
                    vehiculo = Vehiculo.objects.create(
                        id_transportista=transportista,
                        tipo_vehiculo=tipo_vehiculo,
                        capacidad_carga=capacidad_float,
                        placa_vehiculo=placa_vehiculo,
                        estado=estado,
                        documento_propiedad=documento_propiedad if documento_propiedad else None
                    )
                    
                    vehiculos_creados += 1
                    print(f"✅ Vehículo {placa_vehiculo} CREADO exitosamente (ID: {vehiculo.id_vehiculo})")
                    
                except Exception as e:
                    error_msg = f"Fila {fila_num}: Error EXCEPCIÓN - {str(e)}"
                    errores.append(error_msg)
                    print(f"❌ {error_msg}")
                    import traceback
                    traceback.print_exc()
            
            print("\n" + "=" * 50)
            print(f"RESUMEN FINAL:")
            print(f"  - Vehículos creados: {vehiculos_creados}")
            print(f"  - Errores: {len(errores)}")
            for error in errores:
                print(f"    * {error}")
            print("=" * 50)
            
            # Mostrar resultados en el template
            if vehiculos_creados > 0:
                messages.success(request, f'✅ {vehiculos_creados} vehículos cargados exitosamente')
            
            if errores:
                for error in errores[:10]:  # Mostrar primeros 10 errores
                    messages.error(request, error)
                
        except Exception as e:
            print(f"Error general: {str(e)}")
            import traceback
            traceback.print_exc()
            messages.error(request, f'Error al procesar CSV: {str(e)}')
        
        return redirect('mostrar_vehiculos')
    
    return redirect('mostrar_vehiculos')


@login_required
def panel_control(request):
    try:
        usuario_obj = Usuario.objects.get(user=request.user)
        transportista_obj = Transportista.objects.get(id_usuario=usuario_obj)
    except (Usuario.DoesNotExist, Transportista.DoesNotExist):
        messages.error(request, "No tienes perfil de transportista")
        return redirect('inicio_transportista')

    mis_envios_qs = Envio.objects.select_related(
        'id_vehiculo', 'id_compra__id_cliente__id_usuario'
    ).filter(id_transportista=transportista_obj)

    # ── Mini-KPIs ──
    total_asignados = mis_envios_qs.count()
    entregados = mis_envios_qs.filter(estado_envio='Entregado').count()
    pendientes = mis_envios_qs.exclude(
        estado_envio__in=['Entregado', 'Cancelado']
    ).count()

    # ── Fotos pendientes ──
    fotos_pendientes_qs = mis_envios_qs.filter(
        Q(estado_envio='En_Transito', foto_carga__isnull=True) |
        Q(estado_envio='Entregado', foto_carga__isnull=True) |
        Q(estado_envio='Entregado', foto_descarga__isnull=True)
    ).order_by('-id_envio')

    # ── Envíos por vehículo ──
    envios_por_vehiculo = []
    vehiculos_del_tran = Vehiculo.objects.filter(
        id_transportista=transportista_obj
    )
    chart_vehiculo_labels = []
    chart_vehiculo_total = []
    chart_vehiculo_entregados = []
    chart_ingresos_vehiculo_labels = []
    chart_ingresos_vehiculo_data = []

    for v in vehiculos_del_tran:
        envios_vehiculo = mis_envios_qs.filter(id_vehiculo=v)
        total_v = envios_vehiculo.count()
        entregados_v = envios_vehiculo.filter(
            estado_envio='Entregado'
        ).count()
        ingresos_v = envios_vehiculo.filter(
            estado_envio='Entregado'
        ).aggregate(total=Sum('costo_total'))['total'] or 0

        if total_v > 0:
            envios_por_vehiculo.append({
                'vehiculo': v,
                'total': total_v,
                'entregados': entregados_v,
                'ingresos': float(ingresos_v),
            })
            placa = v.placa_vehiculo or f"V-{v.id_vehiculo}"
            chart_vehiculo_labels.append(placa)
            chart_vehiculo_total.append(total_v)
            chart_vehiculo_entregados.append(entregados_v)
            chart_ingresos_vehiculo_labels.append(placa)
            chart_ingresos_vehiculo_data.append(float(ingresos_v))

    # ── Últimos envíos ──
    ultimos_envios = mis_envios_qs.order_by('-id_envio')[:10]

    # ── Charts JSON ──
    chart_data_vehiculo = json.dumps({
        'labels': chart_vehiculo_labels,
        'total': chart_vehiculo_total,
        'entregados': chart_vehiculo_entregados,
    })
    chart_data_ingresos = json.dumps({
        'labels': chart_ingresos_vehiculo_labels,
        'ingresos': chart_ingresos_vehiculo_data,
    })

    return render(request, 'envios/panel_control.html', {
        'total_asignados': total_asignados,
        'entregados': entregados,
        'pendientes': pendientes,
        'fotos_pendientes_count': fotos_pendientes_qs.count(),
        'fotos_pendientes_list': fotos_pendientes_qs,
        'envios_por_vehiculo': envios_por_vehiculo,
        'ultimos_envios': ultimos_envios,
        'chart_data_vehiculo': chart_data_vehiculo,
        'chart_data_ingresos': chart_data_ingresos,
    })


import csv
from django.http import HttpResponse


@login_required
def reporte_envios_csv(request):
    try:
        usuario_obj = Usuario.objects.get(user=request.user)
        transportista_obj = Transportista.objects.get(id_usuario=usuario_obj)
    except (Usuario.DoesNotExist, Transportista.DoesNotExist):
        messages.error(request, "No tienes perfil de transportista")
        return redirect('inicio_transportista')

    envios = Envio.objects.select_related(
        'id_vehiculo', 'id_compra__id_cliente__id_usuario'
    ).filter(id_transportista=transportista_obj).order_by('-id_envio')

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="reporte_envios.csv"'
    response.write('\ufeff')
    writer = csv.writer(response)

    writer.writerow([
        'ID', 'Seguimiento', 'Estado', 'Vehículo', 'Placa',
        'Cliente', 'Origen', 'Destino',
        'Distancia (km)', 'Peso (kg)', 'Costo Total',
        'Fecha Salida', 'Fecha Entrega'
    ])

    for e in envios:
        detalle = e.id_compra.detallescompra_set.first() if e.id_compra else None
        producto = detalle.id_producto if detalle else None
        finca_rel = producto.fincas.first() if producto else None
        finca = finca_rel.id_finca if finca_rel else None

        writer.writerow([
            e.id_envio,
            e.numero_seguimiento or '',
            e.estado_envio or '',
            e.id_vehiculo.tipo_vehiculo if e.id_vehiculo else '',
            e.id_vehiculo.placa_vehiculo if e.id_vehiculo else '',
            f"{e.id_compra.id_cliente.id_usuario.nombre} {e.id_compra.id_cliente.id_usuario.apellido}" if e.id_compra and e.id_compra.id_cliente else '',
            finca.ciudad if finca else '',
            e.id_compra.id_cliente.id_usuario.ciudad if e.id_compra and e.id_compra.id_cliente else '',
            e.distancia_km or 0,
            e.peso_total_kg or 0,
            float(e.costo_total or 0),
            e.fecha_salida or '',
            e.fecha_entrega or '',
        ])

    return response


@login_required
def reporte_ingresos_csv(request):
    try:
        usuario_obj = Usuario.objects.get(user=request.user)
        transportista_obj = Transportista.objects.get(id_usuario=usuario_obj)
    except (Usuario.DoesNotExist, Transportista.DoesNotExist):
        messages.error(request, "No tienes perfil de transportista")
        return redirect('inicio_transportista')

    vehiculos = Vehiculo.objects.filter(id_transportista=transportista_obj)

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="reporte_ingresos.csv"'
    response.write('\ufeff')
    writer = csv.writer(response)

    writer.writerow([
        'Vehículo', 'Placa', 'Tipo', 'Capacidad (kg)',
        'Total Viajes', 'Entregados', 'Ingresos Totales'
    ])

    for v in vehiculos:
        envios_v = Envio.objects.filter(
            id_transportista=transportista_obj, id_vehiculo=v
        )
        total_v = envios_v.count()
        entregados_v = envios_v.filter(estado_envio='Entregado').count()
        ingresos_v = envios_v.filter(
            estado_envio='Entregado'
        ).aggregate(total=Sum('costo_total'))['total'] or 0

        if total_v > 0:
            writer.writerow([
                v.tipo_vehiculo or '',
                v.placa_vehiculo or '',
                v.tipo_vehiculo or '',
                float(v.capacidad_carga or 0),
                total_v,
                entregados_v,
                float(ingresos_v),
            ])

    return response


# ──────────────────────────────────────────────
#  REPORTES EXCEL (openpyxl)
# ──────────────────────────────────────────────

@login_required
def reporte_envios_excel(request):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    try:
        usuario_obj = Usuario.objects.get(user=request.user)
        transportista_obj = Transportista.objects.get(id_usuario=usuario_obj)
    except (Usuario.DoesNotExist, Transportista.DoesNotExist):
        messages.error(request, "No tienes perfil de transportista")
        return redirect('inicio_transportista')

    envios = Envio.objects.select_related(
        'id_vehiculo', 'id_compra__id_cliente__id_usuario'
    ).filter(id_transportista=transportista_obj).order_by('-id_envio')

    wb = openpyxl.Workbook()

    # ── Estilos ──
    hdr_font = Font(bold=True, color='FFFFFF', size=11)
    hdr_fill = PatternFill('solid', fgColor='2f6b31')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    title_font = Font(bold=True, size=14, color='2f6b31')
    kpi_label_font = Font(bold=False, size=11, color='555555')
    kpi_value_font = Font(bold=True, size=16, color='2f6b31')
    kpi_fill = PatternFill('solid', fgColor='F0F7F0')

    # =============== SHEET 1: RESUMEN ===============
    ws_res = wb.active
    ws_res.title = 'Resumen'

    ws_res.merge_cells('A1:F1')
    ws_res['A1'] = 'Reporte de Envíos — Resumen Ejecutivo'
    ws_res['A1'].font = title_font

    total_envios = envios.count()
    entregados = envios.filter(estado_envio='Entregado').count()
    pendientes = envios.exclude(estado_envio__in=['Entregado', 'Cancelado']).count()
    cancelados = envios.filter(estado_envio__in=['Cancelado', 'cancelado']).count()
    ingresos_totales = envios.filter(estado_envio='Entregado').aggregate(
        total=Sum('costo_total'))['total'] or 0
    distancia_total = envios.aggregate(total=Sum('distancia_km'))['total'] or 0
    peso_total = envios.aggregate(total=Sum('peso_total_kg'))['total'] or 0
    promedio_ingreso = float(ingresos_totales / entregados) if entregados > 0 else 0
    tasa_exito = round(entregados / total_envios * 100, 1) if total_envios > 0 else 0

    kpis = [
        ('Total Envíos', total_envios, 'Entregados', entregados),
        ('Pendientes', pendientes, 'Cancelados', cancelados),
        ('Tasa de Éxito', f'{tasa_exito}%', 'Ingresos Totales', f'${float(ingresos_totales):,.0f}'),
        ('Promedio x Viaje', f'${promedio_ingreso:,.0f}', 'Distancia Total', f'{float(distancia_total):,.0f} km'),
        ('Peso Total', f'{float(peso_total):,.0f} kg', '', ''),
    ]

    row = 3
    for c, label in [(1, 'Métrica'), (2, 'Valor'), (3, 'Métrica'), (4, 'Valor')]:
        cell = ws_res.cell(row=row, column=c, value=label)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin_border

    for r, (l1, v1, l2, v2) in enumerate(kpis, start=row + 1):
        ws_res.cell(row=r, column=1, value=l1).font = kpi_label_font
        ws_res.cell(row=r, column=1).border = thin_border
        ws_res.cell(row=r, column=1).fill = kpi_fill
        ws_res.cell(row=r, column=2, value=v1).font = kpi_value_font
        ws_res.cell(row=r, column=2).border = thin_border
        ws_res.cell(row=r, column=2).fill = kpi_fill
        ws_res.cell(row=r, column=3, value=l2).font = kpi_label_font
        ws_res.cell(row=r, column=3).border = thin_border
        ws_res.cell(row=r, column=3).fill = kpi_fill
        ws_res.cell(row=r, column=4, value=v2).font = kpi_value_font if v2 else Font()
        ws_res.cell(row=r, column=4).border = thin_border
        ws_res.cell(row=r, column=4).fill = kpi_fill

    # ── Resumen mensual ──
    monthly = envios.filter(estado_envio='Entregado').annotate(
        mes=TruncMonth('fecha_entrega')
    ).values('mes').annotate(
        total_ingresos=Sum('costo_total'),
        cantidad=Count('id_envio')
    ).order_by('-mes')

    if monthly:
        row += len(kpis) + 2
        ws_res.cell(row=row, column=1, value='Ingresos Mensuales').font = Font(bold=True, size=12, color='2f6b31')
        row += 1
        ws_res.cell(row=row, column=1, value='Mes').font = hdr_font
        ws_res.cell(row=row, column=1).fill = hdr_fill
        ws_res.cell(row=row, column=1).border = thin_border
        ws_res.cell(row=row, column=2, value='Viajes').font = hdr_font
        ws_res.cell(row=row, column=2).fill = hdr_fill
        ws_res.cell(row=row, column=2).border = thin_border
        ws_res.cell(row=row, column=3, value='Ingresos').font = hdr_font
        ws_res.cell(row=row, column=3).fill = hdr_fill
        ws_res.cell(row=row, column=3).border = thin_border
        ws_res.cell(row=row, column=4, value='Promedio').font = hdr_font
        ws_res.cell(row=row, column=4).fill = hdr_fill
        ws_res.cell(row=row, column=4).border = thin_border
        for m in monthly:
            row += 1
            ws_res.cell(row=row, column=1, value=m['mes'].strftime('%B %Y') if m['mes'] else '—').border = thin_border
            ws_res.cell(row=row, column=2, value=m['cantidad']).border = thin_border
            ws_res.cell(row=row, column=3, value=float(m['total_ingresos'] or 0)).border = thin_border
            ws_res.cell(row=row, column=3).number_format = '$#,##0'
            avg = float(m['total_ingresos'] or 0) / m['cantidad'] if m['cantidad'] else 0
            ws_res.cell(row=row, column=4, value=avg).border = thin_border
            ws_res.cell(row=row, column=4).number_format = '$#,##0'

    for col in range(1, 7):
        ws_res.column_dimensions[get_column_letter(col)].width = 22

    # =============== SHEET 2: DETALLE ===============
    ws_det = wb.create_sheet('Detalle')
    headers = [
        'ID', 'Seguimiento', 'Estado', 'Vehículo', 'Placa',
        'Cliente', 'Origen', 'Destino',
        'Dist. (km)', 'Peso (kg)', 'Costo Base', 'Costo Total',
        'Fecha Salida', 'Fecha Entrega', 'Días Tránsito',
        'Costo x km', 'Costo x kg',
    ]
    for c, h in enumerate(headers, 1):
        cell = ws_det.cell(row=1, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin_border

    for i, e in enumerate(envios, start=2):
        detalle = e.id_compra.detallescompra_set.first() if e.id_compra else None
        producto = detalle.id_producto if detalle else None
        finca_rel = producto.fincas.first() if producto else None
        finca = finca_rel.id_finca if finca_rel else None

        dias = None
        if e.fecha_salida and e.fecha_entrega:
            dias = (e.fecha_entrega - e.fecha_salida).days

        costo = float(e.costo_total or 0)
        dist = float(e.distancia_km or 0)
        peso = float(e.peso_total_kg or 0)
        costo_x_km = round(costo / dist, 2) if dist > 0 else 0
        costo_x_kg = round(costo / peso, 2) if peso > 0 else 0

        row_data = [
            e.id_envio,
            e.numero_seguimiento or '',
            e.estado_envio or '',
            e.id_vehiculo.tipo_vehiculo if e.id_vehiculo else '',
            e.id_vehiculo.placa_vehiculo if e.id_vehiculo else '',
            f"{e.id_compra.id_cliente.id_usuario.nombre} {e.id_compra.id_cliente.id_usuario.apellido}" if e.id_compra and e.id_compra.id_cliente else '',
            finca.ciudad if finca else '',
            e.id_compra.id_cliente.id_usuario.ciudad if e.id_compra and e.id_compra.id_cliente else '',
            dist,
            peso,
            float(e.costo_base or 0),
            costo,
            e.fecha_salida or '',
            e.fecha_entrega or '',
            dias or '',
            costo_x_km,
            costo_x_kg,
        ]
        for c, val in enumerate(row_data, 1):
            cell = ws_det.cell(row=i, column=c, value=val)
            cell.border = thin_border
            if isinstance(val, float):
                cell.number_format = '$#,##0' if c in (11, 12, 16, 17) else '#,##0.00'

    for col in range(1, len(headers) + 1):
        ws_det.column_dimensions[get_column_letter(col)].width = 16
    ws_det.column_dimensions[get_column_letter(6)].width = 28
    ws_det.column_dimensions[get_column_letter(2)].width = 18
    ws_det.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{envios.count() + 1}'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_envios.xlsx"'
    wb.save(response)
    return response


@login_required
def reporte_ingresos_excel(request):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    try:
        usuario_obj = Usuario.objects.get(user=request.user)
        transportista_obj = Transportista.objects.get(id_usuario=usuario_obj)
    except (Usuario.DoesNotExist, Transportista.DoesNotExist):
        messages.error(request, "No tienes perfil de transportista")
        return redirect('inicio_transportista')

    vehiculos = Vehiculo.objects.filter(id_transportista=transportista_obj)
    envios_qs = Envio.objects.filter(id_transportista=transportista_obj)

    wb = openpyxl.Workbook()

    hdr_font = Font(bold=True, color='FFFFFF', size=11)
    hdr_fill = PatternFill('solid', fgColor='2f6b31')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    title_font = Font(bold=True, size=14, color='2f6b31')

    # =============== SHEET 1: RESUMEN ===============
    ws_res = wb.active
    ws_res.title = 'Resumen'

    ws_res.merge_cells('A1:E1')
    ws_res['A1'] = 'Reporte de Ingresos — Resumen Ejecutivo'
    ws_res['A1'].font = title_font

    total_ingresos = envios_qs.filter(estado_envio='Entregado').aggregate(
        total=Sum('costo_total'))['total'] or 0
    total_viajes = envios_qs.count()
    entregados = envios_qs.filter(estado_envio='Entregado').count()

    # Totales por vehículo
    data = []
    for v in vehiculos:
        ev = envios_qs.filter(id_vehiculo=v)
        total = ev.count()
        ent = ev.filter(estado_envio='Entregado').count()
        ing = ev.filter(estado_envio='Entregado').aggregate(
            total=Sum('costo_total'))['total'] or 0
        if total > 0:
            avg_ing = float(ing / ent) if ent > 0 else 0
            data.append({
                'placa': v.placa_vehiculo or f'V-{v.id_vehiculo}',
                'tipo': v.tipo_vehiculo or '',
                'capacidad': float(v.capacidad_carga or 0),
                'total': total,
                'entregados': ent,
                'ingresos': float(ing),
                'promedio': avg_ing,
            })

    row = 3
    for c, label in [(1, 'Métrica'), (2, 'Valor'), (3, 'Métrica'), (4, 'Valor')]:
        cell = ws_res.cell(row=row, column=c, value=label)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.border = thin_border

    mejor_vehiculo = max(data, key=lambda x: x['ingresos']) if data else None
    kpis = [
        ('Total Viajes', total_viajes, 'Entregados', entregados),
        ('Ingresos Totales', f'${float(total_ingresos):,.0f}', 'Promedio x Viaje', f'${float(total_ingresos / entregados):,.0f}' if entregados > 0 else '$0'),
        ('Mejor Vehículo', f'{mejor_vehiculo["placa"]} (${mejor_vehiculo["ingresos"]:,.0f})' if mejor_vehiculo else '—', 'Vehículos Activos', len(data)),
    ]
    for r, (l1, v1, l2, v2) in enumerate(kpis, start=row + 1):
        for c, val in [(1, l1), (2, v1), (3, l2), (4, v2)]:
            ws_res.cell(row=r, column=c, value=val).border = thin_border

    # =============== SHEET 2: POR VEHÍCULO ===============
    ws_v = wb.create_sheet('Por Vehículo')
    v_headers = ['Placa', 'Tipo', 'Capacidad (kg)', 'Viajes', 'Entregados', 'Ingresos', 'Promedio x Viaje']
    for c, h in enumerate(v_headers, 1):
        cell = ws_v.cell(row=1, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin_border

    for i, d in enumerate(data, start=2):
        vals = [d['placa'], d['tipo'], d['capacidad'], d['total'], d['entregados'], d['ingresos'], d['promedio']]
        for c, val in enumerate(vals, 1):
            cell = ws_v.cell(row=i, column=c, value=val)
            cell.border = thin_border
            if c in (3,):
                cell.number_format = '#,##0'
            if c in (6, 7):
                cell.number_format = '$#,##0'

    # =============== SHEET 3: MENSUAL ===============
    monthly = envios_qs.filter(estado_envio='Entregado').annotate(
        mes=TruncMonth('fecha_entrega')
    ).values('mes').annotate(
        ingresos=Sum('costo_total'),
        viajes=Count('id_envio'),
    ).order_by('-mes')

    if monthly:
        ws_m = wb.create_sheet('Mensual')
        ws_m.cell(row=1, column=1, value='Mes').font = hdr_font
        ws_m.cell(row=1, column=1).fill = hdr_fill
        ws_m.cell(row=1, column=1).border = thin_border
        ws_m.cell(row=1, column=2, value='Viajes').font = hdr_font
        ws_m.cell(row=1, column=2).fill = hdr_fill
        ws_m.cell(row=1, column=2).border = thin_border
        ws_m.cell(row=1, column=3, value='Ingresos').font = hdr_font
        ws_m.cell(row=1, column=3).fill = hdr_fill
        ws_m.cell(row=1, column=3).border = thin_border
        ws_m.cell(row=1, column=4, value='Promedio').font = hdr_font
        ws_m.cell(row=1, column=4).fill = hdr_fill
        ws_m.cell(row=1, column=4).border = thin_border
        for i, m in enumerate(monthly, start=2):
            avg = float(m['ingresos'] or 0) / m['viajes'] if m['viajes'] else 0
            ws_m.cell(row=i, column=1, value=m['mes'].strftime('%B %Y') if m['mes'] else '—').border = thin_border
            ws_m.cell(row=i, column=2, value=m['viajes']).border = thin_border
            cell = ws_m.cell(row=i, column=3, value=float(m['ingresos'] or 0))
            cell.border = thin_border
            cell.number_format = '$#,##0'
            cell = ws_m.cell(row=i, column=4, value=avg)
            cell.border = thin_border
            cell.number_format = '$#,##0'

    for ws in [ws_res, ws_v]:
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 22

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_ingresos.xlsx"'
    wb.save(response)
    return response


# ──────────────────────────────────────────────
#  REPORTES PDF (reportlab)
# ──────────────────────────────────────────────

@login_required
def reporte_envios_pdf(request):
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.units import inch, mm
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    from io import BytesIO

    try:
        usuario_obj = Usuario.objects.get(user=request.user)
        transportista_obj = Transportista.objects.get(id_usuario=usuario_obj)
    except (Usuario.DoesNotExist, Transportista.DoesNotExist):
        messages.error(request, "No tienes perfil de transportista")
        return redirect('inicio_transportista')

    envios = Envio.objects.select_related(
        'id_vehiculo', 'id_compra__id_cliente__id_usuario'
    ).filter(id_transportista=transportista_obj).order_by('-id_envio')

    total_envios = envios.count()
    entregados = envios.filter(estado_envio='Entregado').count()
    pendientes = envios.exclude(estado_envio__in=['Entregado', 'Cancelado']).count()
    cancelados = envios.filter(estado_envio__in=['Cancelado', 'cancelado']).count()
    ingresos_totales = envios.filter(estado_envio='Entregado').aggregate(
        total=Sum('costo_total'))['total'] or 0
    tasa_exito = round(entregados / total_envios * 100, 1) if total_envios > 0 else 0
    promedio = float(ingresos_totales / entregados) if entregados > 0 else 0

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(letter),
                            leftMargin=20*mm, rightMargin=20*mm,
                            topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    elements = []

    # ── Title ──
    title_style = ParagraphStyle('Title2', parent=styles['Title'],
                                 textColor=colors.HexColor('#2f6b31'),
                                 fontSize=18, spaceAfter=6)
    elements.append(Paragraph('Reporte de Envios — Transportista', title_style))
    elements.append(Paragraph(
        f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}  |  '
        f'{transportista_obj.id_usuario.nombre} {transportista_obj.id_usuario.apellido}',
        styles['Normal']
    ))
    elements.append(Spacer(1, 10))
    elements.append(HRFlowable(width='100%', color=colors.HexColor('#2f6b31'), thickness=2))

    # ── KPI Summary ──
    kpi_style = ParagraphStyle('KPI', parent=styles['Normal'],
                               fontSize=10, alignment=TA_CENTER)
    kpi_data = [
        ['Total Envios', 'Entregados', 'Pendientes', 'Cancelados', 'Tasa Exito', 'Ingresos Totales', 'Promedio x Viaje'],
        [str(total_envios), str(entregados), str(pendientes), str(cancelados),
         f'{tasa_exito}%', f'${float(ingresos_totales):,.0f}', f'${promedio:,.0f}'],
    ]
    kpi_table = Table(kpi_data, colWidths=[80, 80, 80, 80, 70, 100, 100])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2f6b31')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
        ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#f0f7f0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#fafafa')]),
    ]))
    elements.append(Spacer(1, 8))
    elements.append(kpi_table)
    elements.append(Spacer(1, 12))

    # ── Detail Table ──
    hdr_style = ParagraphStyle('Hdr', parent=styles['Normal'],
                               fontSize=7, alignment=TA_CENTER,
                               textColor=colors.white)
    cell_style = ParagraphStyle('Cell', parent=styles['Normal'],
                                fontSize=7, alignment=TA_CENTER)

    table_data = [[Paragraph(h, hdr_style) for h in [
        'ID', 'Seguimiento', 'Estado', 'Vehiculo', 'Placa',
        'Cliente', 'Origen', 'Destino',
        'Dist.', 'Peso', 'Costo', 'Salida', 'Entrega', 'Dias'
    ]]]
    for e in envios:
        detalle = e.id_compra.detallescompra_set.first() if e.id_compra else None
        producto = detalle.id_producto if detalle else None
        finca_rel = producto.fincas.first() if producto else None
        finca = finca_rel.id_finca if finca_rel else None
        dias = ''
        if e.fecha_salida and e.fecha_entrega:
            dias = str((e.fecha_entrega - e.fecha_salida).days)

        table_data.append([Paragraph(str(e.id_envio), cell_style),
                           Paragraph(e.numero_seguimiento or '—', cell_style),
                           Paragraph(e.estado_envio or '—', cell_style),
                           Paragraph(e.id_vehiculo.tipo_vehiculo if e.id_vehiculo else '—', cell_style),
                           Paragraph(e.id_vehiculo.placa_vehiculo if e.id_vehiculo else '—', cell_style),
                           Paragraph(f"{e.id_compra.id_cliente.id_usuario.nombre} {e.id_compra.id_cliente.id_usuario.apellido}" if e.id_compra and e.id_compra.id_cliente else '—', cell_style),
                           Paragraph(finca.ciudad if finca else '—', cell_style),
                           Paragraph(e.id_compra.id_cliente.id_usuario.ciudad if e.id_compra and e.id_compra.id_cliente else '—', cell_style),
                           Paragraph(f'{float(e.distancia_km or 0):.0f}', cell_style),
                           Paragraph(f'{float(e.peso_total_kg or 0):.0f}', cell_style),
                           Paragraph(f'${float(e.costo_total or 0):,.0f}', cell_style),
                           Paragraph(e.fecha_salida.strftime('%d/%m/%Y') if e.fecha_salida else '—', cell_style),
                           Paragraph(e.fecha_entrega.strftime('%d/%m/%Y') if e.fecha_entrega else '—', cell_style),
                           Paragraph(dias, cell_style)])

    col_widths = [25, 65, 55, 55, 50, 95, 50, 55, 40, 40, 65, 60, 60, 35]
    detail_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    detail_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2f6b31')),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#cccccc')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f9f5')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(detail_table)

    doc.build(elements)
    buf.seek(0)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_envios.pdf"'
    response.write(buf.read())
    return response


@login_required
def reporte_ingresos_pdf(request):
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch, mm
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    from io import BytesIO

    try:
        usuario_obj = Usuario.objects.get(user=request.user)
        transportista_obj = Transportista.objects.get(id_usuario=usuario_obj)
    except (Usuario.DoesNotExist, Transportista.DoesNotExist):
        messages.error(request, "No tienes perfil de transportista")
        return redirect('inicio_transportista')

    vehiculos = Vehiculo.objects.filter(id_transportista=transportista_obj)
    envios_qs = Envio.objects.filter(id_transportista=transportista_obj)

    total_viajes = envios_qs.count()
    entregados = envios_qs.filter(estado_envio='Entregado').count()
    ingresos_totales = envios_qs.filter(estado_envio='Entregado').aggregate(
        total=Sum('costo_total'))['total'] or 0
    promedio = float(ingresos_totales / entregados) if entregados > 0 else 0

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter,
                            leftMargin=20*mm, rightMargin=20*mm,
                            topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle('Title2', parent=styles['Title'],
                                 textColor=colors.HexColor('#2f6b31'),
                                 fontSize=18, spaceAfter=6)
    elements.append(Paragraph('Reporte de Ingresos — Transportista', title_style))
    elements.append(Paragraph(
        f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}  |  '
        f'{transportista_obj.id_usuario.nombre} {transportista_obj.id_usuario.apellido}',
        styles['Normal']
    ))
    elements.append(Spacer(1, 10))
    elements.append(HRFlowable(width='100%', color=colors.HexColor('#2f6b31'), thickness=2))

    # KPI summary
    kpi_style = ParagraphStyle('KPI', parent=styles['Normal'],
                               fontSize=11, alignment=TA_CENTER)
    kpi_data = [
        ['Total Viajes', 'Entregados', 'Ingresos Totales', 'Promedio x Viaje'],
        [str(total_viajes), str(entregados),
         f'${float(ingresos_totales):,.0f}', f'${promedio:,.0f}'],
    ]
    kpi_table = Table(kpi_data, colWidths=[110, 90, 130, 130])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2f6b31')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
        ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#f0f7f0')),
    ]))
    elements.append(Spacer(1, 8))
    elements.append(kpi_table)
    elements.append(Spacer(1, 14))

    # Detail by vehicle
    hdr_style = ParagraphStyle('Hdr', parent=styles['Normal'],
                               fontSize=9, alignment=TA_CENTER,
                               textColor=colors.white)
    cell_style = ParagraphStyle('Cell', parent=styles['Normal'],
                                fontSize=9, alignment=TA_CENTER)

    table_data = [[Paragraph(h, hdr_style) for h in [
        'Placa', 'Tipo', 'Capacidad (kg)', 'Viajes', 'Entregados',
        'Ingresos', 'Promedio x Viaje'
    ]]]
    for v in vehiculos:
        ev = envios_qs.filter(id_vehiculo=v)
        total = ev.count()
        ent = ev.filter(estado_envio='Entregado').count()
        ing = ev.filter(estado_envio='Entregado').aggregate(
            total=Sum('costo_total'))['total'] or 0
        if total > 0:
            avg = float(ing / ent) if ent > 0 else 0
            table_data.append([
                Paragraph(v.placa_vehiculo or f'V-{v.id_vehiculo}', cell_style),
                Paragraph(v.tipo_vehiculo or '—', cell_style),
                Paragraph(f'{float(v.capacidad_carga or 0):.0f}', cell_style),
                Paragraph(str(total), cell_style),
                Paragraph(str(ent), cell_style),
                Paragraph(f'${float(ing):,.0f}', cell_style),
                Paragraph(f'${avg:,.0f}', cell_style),
            ])

    cols = [70, 80, 90, 60, 70, 100, 100]
    det_table = Table(table_data, colWidths=cols, repeatRows=1)
    det_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2f6b31')),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#cccccc')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f9f5')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(det_table)

    doc.build(elements)
    buf.seek(0)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_ingresos.pdf"'
    response.write(buf.read())
    return response


def perfil_transportista(request):
    usuario = request.user.usuario
    transportista = Transportista.objects.get(id_usuario=usuario)

    # Vehículos
    vehiculos = Vehiculo.objects.filter(id_transportista=transportista)

    # Envíos
    envios = Envio.objects.filter(id_transportista=transportista)

    # Estadísticas
    entregados = envios.filter(estado_envio="ENTREGADO").count()
    pendientes = envios.exclude(estado_envio="ENTREGADO").count()
    total_envios = envios.count()

    tasa_exito = (entregados / total_envios * 100) if total_envios > 0 else 0

    context = {
        'transportista': transportista,
        'vehiculos': vehiculos,
        'entregados': entregados,
        'pendientes': pendientes,
        'tasa_exito': round(tasa_exito, 1),
    }

    return render(request, 'components/perfil_transportista.html', context)

def editar_perfil_transportista(request):
    usuario = request.user.usuario
    transportista = get_object_or_404(Transportista, id_usuario=usuario)

    if request.method == 'POST':

        #  Usuario
        usuario.nombre = request.POST.get('nombre', usuario.nombre)
        usuario.apellido = request.POST.get('apellido', usuario.apellido)
        usuario.correo = request.POST.get('correo', usuario.correo)
        usuario.telefono = request.POST.get('telefono', usuario.telefono)
        usuario.direccion = request.POST.get('direccion', usuario.direccion)

        if 'foto_perfil' in request.FILES:
            usuario.foto_perfil = request.FILES['foto_perfil']

        # Transportista
        transportista.zonas_entrega = request.POST.get(
            'zonas_entrega',
            transportista.zonas_entrega
        )

        usuario.save()
        transportista.save()

        messages.success(request, "Perfil actualizado correctamente 🚀")

        # IMPORTANTE: nombre de la URL
        return redirect('perfil_transportista')

    return redirect('perfil_transportista')